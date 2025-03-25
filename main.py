import pandas as pd
from openpyxl import load_workbook
import re

def getKeyColumns(codes_string):
    ans = {}
    codes_list = codes_string.split('\n')
    codes = [line.split()[1:] for line in codes_list]
    
    count = 1
    for row in codes:
        # Check if this row has a date (where the last one is '00:00:00')
        if row[-1] == '00:00:00':
            # Remove the time component
            row.pop()
    
    for i in codes:
        #empty string key
        key = ''
        #add all elements of i to key
        for j in i:
            key += j + '-'
        
        #add key to ans
        if key not in ans:
            ans[key] = count
            #print(count, ' : ', key)
        count += 1
    return ans

def getChangedRows(curr, prev):
    ans = []
    for i in curr:
        if i not in prev:
            ans.append([curr[i], i])
    return ans

def clear_filters(file_path):
    try:
        # Load the workbook
        wb = load_workbook(file_path)
        
        # Clear filters from each sheet
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            if ws.auto_filter:
                ws.auto_filter.ref = None
        
        # Save the workbook
        wb.save(file_path)
        print("Filters cleared successfully from all sheets")
    except Exception as e:
        print(f"Error clearing filters: {e}")

def read_excel_file(file_path, sheet_name=0):
    try:
        # Clear filters from all sheets
        clear_filters(file_path)
        
        # Get list of all sheet names
        xl = pd.ExcelFile(file_path)
        print("\nAvailable sheets:", xl.sheet_names)
        
        # Read the new and previous sheets
        current_sheet = len(xl.sheet_names)-2           # 2nd last sheet - New sheet
        previous_current_sheet = len(xl.sheet_names)-3    # 3rd last sheet - Previous sheet
        df_current = pd.read_excel(file_path, sheet_name=current_sheet)
        df_previous = pd.read_excel(file_path, sheet_name=previous_current_sheet)
        
        #Get key columns from new and previous sheets
        print(f"\nReading sheet: {xl.sheet_names[current_sheet]}")
        current_codes = getKeyColumns(df_current[['CUSTOMER CODE','SAP ORDER', 'SAP CODE', 'DESCRIPTION', 'PO', 'PRODUCED QTY', 'INVOICE', 'ESTIMATED DELIVERY DATE']].to_string())
        print('Found',len(current_codes),'rows in new sheet\n---')
        print(f"\nReading sheet: {xl.sheet_names[previous_current_sheet]}")
        previous_codes = getKeyColumns(df_previous[['CUSTOMER CODE', 'SAP ORDER', 'SAP CODE', 'DESCRIPTION', 'PO', 'PRODUCED QTY', 'INVOICE', 'ESTIMATED DELIVERY DATE']].to_string())
        print('Found',len(previous_codes),'rows in previous sheet\n---\n')

        #Get changed rows from new and previous sheets
        changed_rows = getChangedRows(current_codes, previous_codes)
        print('Found',len(changed_rows),'Significant Changes\n---')
    
        # Get Customer Order and SAP Order from changed rows
        ans = []
        for i in changed_rows:
            keyCodes = i[1].split('-')
            ans.append('-'.join(keyCodes[:2]))
        
        print('Customer Order-SAP Order')
        for i in ans:
            print(i)
        return df_current

    except Exception as e:
        print(f"Error reading Excel file: {e}")
        return None

if __name__ == "__main__":
    excel_file_path = "test1.xlsx"
    # Read first sheet (by index)
    print("Reading first sheet:")
    df1 = read_excel_file(excel_file_path, sheet_name=0)